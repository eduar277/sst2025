from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .forms import RiskEvaluationForm, FiltroRiesgosForm
from .models import RiskEvaluation
from users.utils import role_required
from django.contrib import messages
from django.db.models import Count
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

@login_required
def evaluar_riesgo(request):
    if request.method == 'POST':
        form = RiskEvaluationForm(request.POST, request.FILES)
        if form.is_valid():
            riesgo = form.save(commit=False)
            riesgo.usuario = request.user
            riesgo.save()
            return redirect('mis_riesgos')
    else:
        form = RiskEvaluationForm()
    return render(request, 'risks/evaluar.html', {'form': form})

@login_required
def mis_riesgos(request):
    riesgos = RiskEvaluation.objects.filter(usuario=request.user).order_by('-fecha')
    return render(request, 'risks/mis_riesgos.html', {'riesgos': riesgos})

@login_required
@role_required(['admin', 'empleador'])
def listar_todos_los_riesgos(request):
    riesgos = RiskEvaluation.objects.all().order_by('-fecha')
    form = FiltroRiesgosForm(request.GET or None)

    if form.is_valid():
        if form.cleaned_data['empleado']:
            riesgos = riesgos.filter(usuario=form.cleaned_data['empleado'])
        if form.cleaned_data['factor']:
            riesgos = riesgos.filter(factor=form.cleaned_data['factor'])
        if form.cleaned_data['fecha_inicio']:
            riesgos = riesgos.filter(fecha__date__gte=form.cleaned_data['fecha_inicio'])
        if form.cleaned_data['fecha_fin']:
            riesgos = riesgos.filter(fecha__date__lte=form.cleaned_data['fecha_fin'])

    return render(request, 'risks/todos_riesgos.html', {
        'riesgos': riesgos,
        'form': form,
    })

@login_required
@role_required(['admin', 'empleador'])
def actualizar_estado_riesgo(request, riesgo_id):
    riesgo = get_object_or_404(RiskEvaluation, id=riesgo_id)
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in dict(RiskEvaluation.ESTADOS):
            riesgo.estado = nuevo_estado
            riesgo.save()
            messages.success(request, f"Estado actualizado a '{riesgo.get_estado_display()}' para {riesgo.usuario.get_full_name()}")
    return redirect('todos_riesgos')

@login_required
@role_required(['admin', 'empleador'])
def exportar_riesgos_pdf(request):
    from django.template.loader import get_template
    from xhtml2pdf import pisa

    riesgos = RiskEvaluation.objects.all().order_by('-fecha')
    form = FiltroRiesgosForm(request.GET or None)

    if form.is_valid():
        if form.cleaned_data['empleado']:
            riesgos = riesgos.filter(usuario=form.cleaned_data['empleado'])
        if form.cleaned_data['factor']:
            riesgos = riesgos.filter(factor=form.cleaned_data['factor'])
        if form.cleaned_data['fecha_inicio']:
            riesgos = riesgos.filter(fecha__date__gte=form.cleaned_data['fecha_inicio'])
        if form.cleaned_data['fecha_fin']:
            riesgos = riesgos.filter(fecha__date__lte=form.cleaned_data['fecha_fin'])

    template = get_template('risks/reporte_pdf.html')
    html = template.render({'riesgos': riesgos})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_riesgos.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

@login_required
@role_required(['admin', 'empleador'])
def exportar_riesgos_excel(request):
    riesgos = RiskEvaluation.objects.all().order_by('-fecha')
    form = FiltroRiesgosForm(request.GET or None)

    if form.is_valid():
        if form.cleaned_data['empleado']:
            riesgos = riesgos.filter(usuario=form.cleaned_data['empleado'])
        if form.cleaned_data['factor']:
            riesgos = riesgos.filter(factor=form.cleaned_data['factor'])
        if form.cleaned_data['fecha_inicio']:
            riesgos = riesgos.filter(fecha__date__gte=form.cleaned_data['fecha_inicio'])
        if form.cleaned_data['fecha_fin']:
            riesgos = riesgos.filter(fecha__date__lte=form.cleaned_data['fecha_fin'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riesgos Reportados"
    columnas = ['Empleado', 'Factor', 'Descripción', 'Fecha', 'Estado']
    ws.append(columnas)

    for r in riesgos:
        ws.append([
            r.usuario.get_full_name(),
            r.get_factor_display(),
            r.descripcion,
            r.fecha.strftime('%Y-%m-%d %H:%M'),
            r.get_estado_display()
        ])

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="riesgos_reportados.xlsx"'
    wb.save(response)
    return response
